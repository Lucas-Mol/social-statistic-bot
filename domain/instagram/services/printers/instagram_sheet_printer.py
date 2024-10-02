from domain.instagram.objects.instagram_profile import Instagram_Profile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

FILE_NAME = '/instagram_statistics.xlsx'

def print(instagram_profile: Instagram_Profile, path):
    file_path = path + FILE_NAME
    
    df_profile = _get_df_profile(instagram_profile)
    df_posts = _get_df_media(instagram_profile)

    _create_excel_file(df_profile, df_posts, file_path)

    wb = load_workbook(file_path)
    ws_profile = wb['Tabela Perfil']
    ws_posts = wb['Tabela Postagens']

    _adjust_cell_size(ws_profile)
    _adjust_cell_size_posts(ws_posts)

    _color_fill_playlist_cells(ws_profile)
    _color_fill_playlist_cells(ws_posts)

    wb.save(file_path)

def _get_df_profile(instagram_profile: Instagram_Profile):
    data = {
        '': ['Perfil'],
        'Nome': [str(instagram_profile.username)],
        'Seguidores': [str(instagram_profile.followers_count)],
        'Curtidas': [str(instagram_profile.total_likes)],
        'Postagens': [str(instagram_profile.media_count)]
    }

    return pd.DataFrame(data)

def _get_df_media(instagram_profile: Instagram_Profile):
    data = {
        'Nome': _get_name_column_lines(instagram_profile),
        'Curtidas': _get_likes_column_lines(instagram_profile),
        'Comentários': _get_comment_column_lines(instagram_profile)
    }

    return pd.DataFrame(data)

def _get_name_column_lines(instagram_profile: Instagram_Profile):
    lines = []

    for post in instagram_profile.posts:
        lines.append(str(post.name))

    return lines

def _get_likes_column_lines(instagram_profile: Instagram_Profile):
    lines = []

    for post in instagram_profile.posts:
        lines.append(str(post.likes))

    return lines


def _get_comment_column_lines(instagram_profile: Instagram_Profile):
    lines = []

    for post in instagram_profile.posts:
        lines.append(str(post.comments))

    return lines

def _create_excel_file(df_profile, df_posts, file_path):
    with pd.ExcelWriter(file_path) as writer:
        df_profile.to_excel(writer, sheet_name='Tabela Perfil', index=False)
        df_posts.to_excel(writer, sheet_name='Tabela Postagens', index=False)

def _adjust_cell_size(ws):
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) 
        ws.column_dimensions[column[0].column_letter].width = adjusted_width  

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20

def _adjust_cell_size_posts(ws):
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if 1 == cell.column:
                    max_length = len(str(cell.value)) / 5
                elif len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) 
        ws.column_dimensions[column[0].column_letter].width = adjusted_width  

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, shrink_to_fit=True)
        ws.row_dimensions[row[0].row].height = 30

def _color_fill_playlist_cells(ws_posts):
    fill = PatternFill(start_color="86B8D5", end_color="86B8D5", fill_type="solid")

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws_posts.iter_rows(min_row=1, max_row=1): 
        for cell in row:  
            cell.fill = fill
            cell.border = thin_border

    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_odd = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row in ws_posts.iter_rows(min_row=2, max_row=ws_posts.max_row):
        row_number = row[0].row
        fill = fill_even if row_number % 2 == 0 else fill_odd
        for cell in row:
            cell.fill = fill
            cell.border = thin_border