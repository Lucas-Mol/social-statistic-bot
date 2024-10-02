from domain.youtube.objects.youtube_channel import Youtube_Channel

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

FILE_NAME = '/youtube_statistics.xlsx'

def print(youtube_channel: Youtube_Channel, path):
    file_path = path + FILE_NAME
    
    df_channel = _get_df_channel(youtube_channel)
    df_playlist = _get_df_videos(youtube_channel)

    _create_excel_file(df_channel, df_playlist, file_path)

    wb = load_workbook(file_path)
    ws_channel = wb['Tabela Canal']
    ws_videos = wb['Tabela Vídeos']

    _adjust_cell_size(ws_channel)
    _adjust_cell_size(ws_videos)

    _color_fill_playlist_cells(ws_channel)
    _color_fill_playlist_cells(ws_videos)

    wb.save(file_path)

def _get_df_channel(youtube_channel: Youtube_Channel):
    data = {
        '': ['Canal'],
        'Nome': [str(youtube_channel.username)],
        'Inscritos': [str(youtube_channel.subs_count)],
        'Visualizações': [str(youtube_channel.view_count)],
        'Vídeos': [str(youtube_channel.video_count)]
    }

    return pd.DataFrame(data)

def _get_df_videos(youtube_channel: Youtube_Channel):
    data = {
        '': _get_first_column_lines(youtube_channel),
        'Nome': _get_name_column_lines(youtube_channel),
        'Curtidas': _get_likes_column_lines(youtube_channel),
        'Visualizações': _get_views_column_lines(youtube_channel),
        'Comentários': _get_comment_column_lines(youtube_channel)
    }

    return pd.DataFrame(data)

def _get_first_column_lines(youtube_channel: Youtube_Channel):
    lines = []

    for playlist in youtube_channel.playlists:
        lines.append('Playlist')
        for video in playlist.videos:
            lines.append('Vídeo')

    return lines

def _get_name_column_lines(youtube_channel: Youtube_Channel):
    lines = []

    for playlist in youtube_channel.playlists:
        lines.append(str(playlist.name))
        for video in playlist.videos:
            lines.append(str(video.name))

    return lines

def _get_likes_column_lines(youtube_channel: Youtube_Channel):
    lines = []

    for playlist in youtube_channel.playlists:
        lines.append(str(playlist.likes))
        for video in playlist.videos:
            lines.append(str(video.likes))

    return lines

def _get_views_column_lines(youtube_channel: Youtube_Channel):
    lines = []

    for playlist in youtube_channel.playlists:
        lines.append(str(playlist.views))
        for video in playlist.videos:
            lines.append(str(video.views))

    return lines

def _get_comment_column_lines(youtube_channel: Youtube_Channel):
    lines = []

    for playlist in youtube_channel.playlists:
        lines.append(str(playlist.comments))
        for video in playlist.videos:
            lines.append(str(video.comments))

    return lines

def _create_excel_file(df_channel, df_playlist, file_path):
    with pd.ExcelWriter(file_path) as writer:
        df_channel.to_excel(writer, sheet_name='Tabela Canal', index=False)
        df_playlist.to_excel(writer, sheet_name='Tabela Vídeos', index=False)

def _color_fill_playlist_cells(ws_videos):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_odd = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row in ws_videos.iter_rows(min_row=2, max_row=ws_videos.max_row):
        row_number = row[0].row
        fill = fill_even if row_number % 2 == 0 else fill_odd
        for cell in row:
            cell.fill = fill
            cell.border = thin_border

    header_fill = PatternFill(start_color="86B8D5", end_color="86B8D5", fill_type="solid")

    for row in ws_videos.iter_rows(min_row=1, max_row=1): 
        for cell in row:  
            cell.fill = header_fill
            cell.border = thin_border

    playlist_fill = PatternFill(start_color="82C7F5", end_color="82C7F5", fill_type="solid")
    
    for row in ws_videos.iter_rows(min_row=2, max_row=ws_videos.max_row): 
        if row[0].value == "Playlist":  
            for cell in row:  
                cell.fill = playlist_fill
                cell.border = thin_border

def _adjust_cell_size(ws):
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) 
        ws.column_dimensions[column[0].column_letter].width = adjusted_width  

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20